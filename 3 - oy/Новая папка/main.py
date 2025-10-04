import openpyxl
path = "excel.xlsx"
wb = openpyxl.load_workbook(path)
sheet = wb.active
max_row = sheet.max_row
max_col = sheet.max_column

list_key = [
    sheet.cell(row=1, column=i).value
    for i in range(1, max_col + 1)
]

list_value = [
    [sheet.cell(row=i, column=j).value for j in range(1, max_col + 1)]
    for i in range(2, max_row + 1)
]


class Avto:
    def __init__(self, name, number, color, company):
        self.name = name
        self.number = number
        self.color = color
        self.company = company
    def info(self):
        return f"name: {self.name}, number: {self.number}, color: {self.color}, company: {self.company}"

avtos = [
    Avto(name=list_value[i][0],
          number=list_value[i][1],
            color=list_value[i][2],
              company=list_value[i][3])
              
    for i in range(len(list_value))
]

for d in avtos:
    print(d.info)

